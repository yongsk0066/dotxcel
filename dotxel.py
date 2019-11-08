import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
from PIL import Image
import numpy as np
import cv2
import math


# wb = openpyxl.load_workbook('ex.xlsx')
# ws = wb.active

# img = Image.open(src).convert('RGB')

# size = int(input("size >>"))
def paint_exel(src, K=6, size = 60):
    wb = openpyxl.Workbook() 
    # Sheet_name = wb.sheetnames
    ws = wb.get_sheet_by_name("Sheet")
    # ws = wb.create_sheet("Sheet1")
    print(ws)
    img = Image.open(src).convert('RGB')
    width = math.floor(size * img.size[0]/img.size[1])

    resize_img = img.resize((width, size))
    img_arr = np.asarray(resize_img)

    Z = img_arr.reshape((-1,3))

    # convert to np.float32
    Z = np.float32(Z)

    # define criteria, number of clusters(K) and apply kmeans()
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 10, 1.0)
    K = K
    ret, label, center = cv2.kmeans(Z, K, None, criteria, 10, cv2.KMEANS_RANDOM_CENTERS)

    # Now convert back into uint8, and make original image
    center = np.uint8(center)
    res = center[label.flatten()]
    res2 = res.reshape((size,width,3))

    # cv2.imshow('res2',res2)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()

    for i in range(1, size + 1):
        ws.row_dimensions[i].height = 5
        for j in range(1, width + 1):
            color = '{:02x}{:02x}{:02x}'.format(res2[i - 1][j - 1][0],
                                                res2[i - 1][j - 1][1],
                                                res2[i - 1][j - 1][2])
            if color != "000000":
                targetFill = PatternFill(start_color=color,
                                        end_color=color,
                                        fill_type='solid')
                ws.cell(row=i, column=j).fill = targetFill
            ws.column_dimensions[get_column_letter(j)].width = 1

    wb.save("output.xlsx")
    wb.close()
